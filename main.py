import tkinter.messagebox
import tkinter as tk

import openpyxl
import pymysql.cursors
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import teat
import DataBase
import datetime
import pandas as pd



class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.db_name = None
        self.table_name = None
        self.title("Создание базы данных")

        self.db_label = tk.Label(self, text="Имя базы данных:")
        self.db_label.pack()
        self.db_entry = tk.Entry(self, textvariable=tk.StringVar())
        self.db_entry.pack()

        self.table_label = tk.Label(self, text="Имя таблицы:")
        self.table_label.pack()
        self.table_entry = tk.Entry(self, textvariable=tk.StringVar())
        self.table_entry.pack()

        self.fio = tk.Label(self, text="ФИО:")
        self.fio.pack()
        self.fio = tk.Entry(self, textvariable=tk.StringVar())
        self.fio.pack()

        self.birthday = tk.Label(self, text="День рождения (dd:mm:yyyy):")
        self.birthday.pack()
        self.birthday = tk.Entry(self, textvariable=tk.StringVar())
        self.birthday.pack()

        self.file1_label1 = tk.Label(self, text="Имя файла эксель:")
        self.file1_label1.pack()
        self.file1_entry1 = tk.Entry(self, textvariable=tk.StringVar())
        self.file1_entry1.pack()

        self.db_label.pack(padx=50, pady=2)
        self.table_label.pack(padx=50, pady=2)
        self.fio.pack(padx=50, pady=2)
        self.birthday.pack(padx=50, pady=2)
        self.file1_label1.pack(padx=50, pady=2)

    def create_database(self):
        self.db_name = self.db_entry.get()
        self.table_name = self.table_entry.get()

        db = DataBase.DataBase(self.db_name, self.table_name)
        db.check_db()
        db.check_table()

    def brth_days_count(self, birth):
        days_since_birthday = teat.brth_days_count(birth)
        return days_since_birthday

    def save_result(self):
        fio = self.fio.get()
        birthday = self.birthday.get()
        days_count = self.brth_days_count(birthday)


        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()

        try:
            with connection.cursor() as cursor:
                cursor.execute(f"INSERT INTO {db1.name_tb} (fio, birthday,days_count) VALUES (%s, %s, %s)",
                               (fio, birthday, days_count))
                connection.commit()

                cursor.execute(f"SELECT * FROM {db1.name_tb}")
                print(cursor.fetchall()[-1])

        except pymysql.err.DataError as e:
            print('Ошибка с данными:', e)

        except pymysql.err.DatabaseError as e:
            print(e)

    def list_tb(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        cursor = connection.cursor()
        tb_in_db = "SHOW TABLES;"
        cursor.execute(tb_in_db)
        tables = cursor.fetchall()

        table_list = [table[0] for table in tables]
        table_list_str = "\n".join(table_list)

        tkinter.messagebox.showinfo("Список таблиц", table_list_str)

    def save_to_excel(self):
        db1 = DataBase.DataBase(self.db_name, self.table_name)
        connection = db1.con_db()
        try:
            new_df = pd.read_sql("SELECT * FROM " + self.table_name, connection)
            wb = openpyxl.Workbook()
            ws = wb.active

            for r in dataframe_to_rows(new_df, index=False, header=True):
                ws.append(r)

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            file1 = self.file1_entry1.get()
            wb.save(file1)
            print(new_df)

            tkinter.messagebox.showinfo("Импорт в эксель", file1)

        except pymysql.err.DatabaseError as e:
            print(e)
        return


app = App()

create_button1 = tk.Button(app, text="Создать БД", command=app.create_database)
create_button1.pack()

create_button = tk.Button(app, text="Создать запись", command=app.save_result)
create_button.pack()

list_button = tk.Button(app, text="Показать список таблиц", command=app.list_tb)
list_button.pack()

excel_button = tk.Button(app, text="Импорт в эксель", command=app.save_to_excel)
excel_button.pack()


app.mainloop()
